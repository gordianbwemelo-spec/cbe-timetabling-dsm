"""Build the printable Excel workbook for a semester, in memory (BytesIO)."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
STARTS = [7, 9, 11, 13, 15, 17, 19]
PERIODS = [f"{t:02d}:00-{t+2:02d}:00" for t in STARTS]
EVE = {17, 19}
FONT = "Arial"; NAVY = "16305C"; BLUE = "2E5496"; LBLUE = "DCE6F7"; RED = "FBE0DD"; AMBER = "FDF1D6"; GREEN = "DFF0E4"; VAC = "F6F7FB"; GREY = "E4E7EE"
def Fn(**k): return Font(name=FONT, **k)
thin = Side(style="thin", color="C4CCDA"); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
def time_of(t): return f"{t:02d}:00-{t+2:02d}:00"

def hdr(ws, row, cols, fill=BLUE):
    for j, v in enumerate(cols, 1):
        c = ws.cell(row, j, v); c.font = Fn(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=fill); c.alignment = CEN; c.border = BORD
def title(ws, text, sub, ncol):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    c = ws.cell(1, 1, text); c.font = Fn(bold=True, size=14, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CEN
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    c = ws.cell(2, 1, sub); c.font = Fn(italic=True, size=9, color=NAVY); c.alignment = CEN; ws.row_dimensions[1].height = 24

def build_workbook(sem, S, VEN, d, model_note):
    Vmap = {v["venue"]: v for v in VEN}; m = d["metrics"]
    wb = Workbook()
    ws = wb.active; ws.title = "Overview"
    title(ws, f"COLLEGE OF BUSINESS EDUCATION — Master Timetable, Semester {sem}", "Dashboard — figures identical to the web application", 4)
    est = " (occupancy estimated at source)" if m["estimated"] else ""
    rows = [("Class sessions", m["sessions"]), ("Venues", m["venues"]), ("Seat capacity / period", m["seatcap"]),
            ("Venue-periods available", m["avail"]), ("Venue-periods used", m["used"]), ("Vacant venue-periods", m["vacant"]),
            ("Overall utilisation", f'{m["util"]}%'), ("Peak-period utilisation", f'{m["peak"]}%'),
            ("Instructors", m["instructors"]), ("Modules" + est, m["modules"]),
            ("Hard rule-breaks", m["hard"]), ("Items to review", m["review"]),
            ("Instructors over a cap", m["overloads"]), ("At/over soft limit", m["softs"])]
    ws.cell(4, 1, "METRIC").font = Fn(bold=True, color="FFFFFF"); ws.cell(4, 1).fill = PatternFill("solid", fgColor=BLUE); ws.cell(4, 1).border = BORD
    ws.cell(4, 2, "VALUE").font = Fn(bold=True, color="FFFFFF"); ws.cell(4, 2).fill = PatternFill("solid", fgColor=BLUE); ws.cell(4, 2).border = BORD
    for i, (k, v) in enumerate(rows, 1):
        a = ws.cell(4 + i, 1, k); b = ws.cell(4 + i, 2, v); a.font = Fn(); b.font = Fn(bold=True); a.border = BORD; b.border = BORD; b.alignment = Alignment(horizontal="right")
        if k in ("Hard rule-breaks", "Instructors over a cap") and v: b.fill = PatternFill("solid", fgColor=RED)
    busy = sorted(d["vutil"], key=lambda v: -v["periods_used"])[:12]
    for col, lab in [(4, "BUSIEST VENUES"), (5, "Used"), (6, "Util%")]:
        c = ws.cell(4, col, lab); c.font = Fn(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE); c.border = BORD
    for i, v in enumerate(busy, 1):
        ws.cell(4 + i, 4, f'{v["venue"]} [{v["capacity"]}]').font = Fn(); ws.cell(4 + i, 4).border = BORD
        ws.cell(4 + i, 5, v["periods_used"]).font = Fn(); ws.cell(4 + i, 5).border = BORD
        ws.cell(4 + i, 6, v["utilisation"]).font = Fn(); ws.cell(4 + i, 6).border = BORD
    tb = 4 + len(busy) + 2
    for col, lab in [(4, "HIGHEST INSTRUCTOR LOADS"), (5, "Tot h"), (6, "Flag")]:
        c = ws.cell(tb, col, lab); c.font = Fn(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=BLUE); c.border = BORD
    for i, w in enumerate(d["workload"][:12], 1):
        ws.cell(tb + i, 4, w["instructor"]).font = Fn(); ws.cell(tb + i, 4).border = BORD
        ws.cell(tb + i, 5, w["total_h"]).font = Fn(); ws.cell(tb + i, 5).border = BORD
        cc = ws.cell(tb + i, 6, ", ".join(w["flags"])); cc.font = Fn(); cc.border = BORD
        if w["flags"]: cc.fill = PatternFill("solid", fgColor=AMBER)
    nrow = tb + 14
    n = ws.cell(nrow, 1, "Note: " + (model_note or "")); n.font = Fn(italic=True, size=8); n.alignment = LEFT
    ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow + 2, end_column=6)
    for col, wd in {"A": 32, "B": 14, "C": 3, "D": 30, "E": 8, "F": 18}.items(): ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A4"
    # Venue Capacity
    ws = wb.create_sheet("Venue Capacity"); title(ws, f"Venue Capacity — Semester {sem}", "Seat capacity, type and premises", 4)
    hdr(ws, 4, ["Venue", "Capacity", "Type", "Premises"]); r = 5
    for v in sorted(VEN, key=lambda v: (v["premises"], -v["capacity"])):
        for j, val in enumerate([v["venue"], v["capacity"], v["type"], v["premises"]], 1):
            c = ws.cell(r, j, val); c.font = Fn(); c.border = BORD
        r += 1
    ws.cell(r, 1, "TOTAL SEAT CAPACITY / SESSION").font = Fn(bold=True); ws.cell(r, 1).border = BORD
    ws.cell(r, 2, sum(v["capacity"] for v in VEN)).font = Fn(bold=True); ws.cell(r, 2).border = BORD; ws.cell(r, 2).fill = PatternFill("solid", fgColor=LBLUE)
    for col, wd in {"A": 24, "B": 11, "C": 13, "D": 11}.items(): ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:D{r-1}"
    # Master List
    ws = wb.create_sheet("Master List"); title(ws, f"Master List — Semester {sem}", "One filterable row per session", 12)
    hdr(ws, 4, ["Day", "Period", "Venue", "Cap", "Premises", "Cohort", "NTA", "Stream", "Module", "Code", "Instructor", "Occupancy"]); r = 5
    for s in sorted(S, key=lambda s: (DAYS.index(s["day"]), s["t"], s["venue"])):
        v = Vmap.get(s["venue"], {})
        vals = [s["day"], time_of(s["t"]), s["venue"], v.get("capacity", s["cap"]), v.get("premises", ""), s["prog"], s["nta"], s["stream"], s["mod"], s["code"], s["instr"], s["occ"]]
        for j, val in enumerate(vals, 1):
            c = ws.cell(r, j, val); c.font = Fn(size=9); c.border = BORD; c.alignment = LEFT
            if j == 12 and v and s["occ"] > v.get("capacity", 9999) + 10: c.fill = PatternFill("solid", fgColor=RED)
        r += 1
    for i, wd in enumerate([9, 11, 20, 6, 9, 22, 9, 6, 28, 11, 24, 9], 1): ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:L{r-1}"
    # Day grids
    grid = defaultdict(dict)
    for s in S: grid[(s["day"], s["venue"])][s["t"]] = f'{s["prog"]}\n{s["mod"]}\n{s["instr"]} ({s["occ"]})'
    vlist = sorted(VEN, key=lambda v: (0 if v["premises"] == "Main" else 1, -v["capacity"]))
    for day in DAYS:
        ws = wb.create_sheet(day); title(ws, f"{day} — Venue × Period ({sem})", "shaded = vacant; grey = no session; footer = used / seats", 8)
        hdr(ws, 4, ["VENUE [cap]"] + PERIODS); r = 5
        for v in vlist:
            ws.cell(r, 1, f'{v["venue"]} [{v["capacity"]}] {v["premises"]}').font = Fn(bold=True, size=8); ws.cell(r, 1).border = BORD; ws.cell(r, 1).alignment = LEFT
            for j, t in enumerate(STARTS, 2):
                c = ws.cell(r, j); c.border = BORD; c.font = Fn(size=8); c.alignment = CEN
                if v["premises"] == "Saba" and t in EVE: c.value = "—"; c.fill = PatternFill("solid", fgColor=GREY)
                elif t in grid[(day, v["venue"])]: c.value = grid[(day, v["venue"])][t]; c.fill = PatternFill("solid", fgColor=GREEN)
                else: c.fill = PatternFill("solid", fgColor=VAC)
            ws.row_dimensions[r].height = 32; r += 1
        ws.cell(r, 1, "VENUES USED").font = Fn(bold=True); ws.cell(r, 1).border = BORD
        ws.cell(r + 1, 1, "SEATS FILLED").font = Fn(bold=True); ws.cell(r + 1, 1).border = BORD
        for j, t in enumerate(STARTS, 2):
            vu = sum(1 for s in S if s["day"] == day and s["t"] == t); sf = sum(s["occ"] for s in S if s["day"] == day and s["t"] == t)
            a = ws.cell(r, j, vu); a.font = Fn(bold=True); a.border = BORD; a.alignment = CEN; a.fill = PatternFill("solid", fgColor=LBLUE)
            b = ws.cell(r + 1, j, sf); b.font = Fn(bold=True); b.border = BORD; b.alignment = CEN; b.fill = PatternFill("solid", fgColor=LBLUE)
        ws.column_dimensions["A"].width = 22
        for j in range(2, 9): ws.column_dimensions[get_column_letter(j)].width = 17
        ws.freeze_panes = "B5"
    # Instructor TT
    ws = wb.create_sheet("Instructor TT"); title(ws, f"Instructor Timetables — Semester {sem}", "Weekly grid per lecturer", 8); r = 4
    itt = defaultdict(dict)
    for s in S:
        if s["instr"]: itt[s["instr"]][(s["day"], s["t"])] = f'{s["mod"]} · {s["prog"]} @{s["venue"]}'
    for name in sorted(itt):
        ws.cell(r, 1, name).font = Fn(bold=True, color="FFFFFF"); ws.cell(r, 1).fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 1
        hdr(ws, r, ["Day"] + PERIODS); r += 1
        for day in DAYS:
            ws.cell(r, 1, day).font = Fn(bold=True); ws.cell(r, 1).border = BORD
            for j, t in enumerate(STARTS, 2):
                c = ws.cell(r, j); c.border = BORD; c.font = Fn(size=8); c.alignment = CEN
                if (day, t) in itt[name]: c.value = itt[name][(day, t)]; c.fill = PatternFill("solid", fgColor=GREEN)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 14
    for j in range(2, 9): ws.column_dimensions[get_column_letter(j)].width = 18
    # Workload
    ws = wb.create_sheet("Workload"); title(ws, f"Instructor Workload — Semester {sem}", "Caps 7 / 32h / 20h (soft 6 / 28 / 16)", 7)
    hdr(ws, 4, ["Instructor", "Department", "Modules", "Daytime h", "Evening h", "Total h", "Flag"]); r = 5
    for w in d["workload"]:
        over = any(f.isupper() and f != f.lower() for f in w["flags"])
        for j, val in enumerate([w["instructor"], w["dept"] or "—", w["modules"], w["daytime_h"], w["evening_h"], w["total_h"], ", ".join(w["flags"])], 1):
            c = ws.cell(r, j, val); c.font = Fn(size=9); c.border = BORD
            if j == 7 and w["flags"]: c.fill = PatternFill("solid", fgColor=(RED if over else AMBER))
        r += 1
    for i, wd in enumerate([24, 28, 9, 10, 10, 8, 22], 1): ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:G{r-1}"
    # Venue Utilisation
    ws = wb.create_sheet("Venue Utilisation"); title(ws, f"Venue Utilisation — Semester {sem}", "Periods & seat-periods used vs available", 7)
    hdr(ws, 4, ["Venue", "Cap", "Premises", "Periods used", "Periods avail", "Seat-periods used", "Utilisation%"]); r = 5
    for v in sorted(d["vutil"], key=lambda v: (v["premises"], -v["periods_used"])):
        for j, val in enumerate([v["venue"], v["capacity"], v["premises"], v["periods_used"], v["periods_avail"], v["seat_periods_used"], v["utilisation"]], 1):
            c = ws.cell(r, j, val); c.font = Fn(size=9); c.border = BORD
        r += 1
    for i, wd in enumerate([20, 7, 10, 13, 13, 17, 13], 1): ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:G{r-1}"
    # Red-flags
    ws = wb.create_sheet("Red-flags"); title(ws, f"Red-flags — Semester {sem}", "Hard rule-breaks and items to review", 3)
    n = ws.cell(3, 1, model_note or ""); n.font = Fn(italic=True, size=9); n.alignment = LEFT; ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    hdr(ws, 5, ["Severity", "Rule", "Detail"]); r = 6
    order = {"hard": 0, "review": 1}
    for f in sorted(d["flags"], key=lambda f: (order[f["severity"]], f["type"])):
        a = ws.cell(r, 1, f["severity"].upper()); b = ws.cell(r, 2, f["type"]); cc = ws.cell(r, 3, f["detail"])
        for x in (a, b, cc): x.font = Fn(size=9); x.border = BORD; x.alignment = LEFT
        a.fill = PatternFill("solid", fgColor=(RED if f["severity"] == "hard" else AMBER)); a.font = Fn(size=9, bold=True)
        r += 1
    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 22; ws.column_dimensions["C"].width = 95
    ws.freeze_panes = "A6"; ws.auto_filter.ref = f"A5:C{r-1}"
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio
